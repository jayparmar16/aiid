from __future__ import annotations

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .config import PipelineConfig, ThemeConfig


# ---------------------------------------------------------------------------
# Column-group resolution (config-driven)
# ---------------------------------------------------------------------------

def _col_group(col: str, config: PipelineConfig) -> tuple[str, str]:
    """Return the grouping name and hex colour for a given column based on config."""
    for style_name, style_conf in config.styles.items():
        if col in style_conf.columns:
            return (style_name, style_conf.color)
    for tax_name, tax_config in config.taxonomies.items():
        if col in tax_config.mapping.values():
            return (tax_name, tax_config.color)
    return ("Other", config.styles["Other"].color)


def _apply_order(df: pd.DataFrame, preferred: list[str]) -> pd.DataFrame:
    """Reorder columns: preferred order first, then any remaining columns."""
    order = [c for c in preferred if c in df.columns]
    order += [c for c in df.columns if c not in order]
    return df[order]


# ---------------------------------------------------------------------------
# Shared small helpers
# ---------------------------------------------------------------------------

def _thin_border(theme: ThemeConfig) -> Border:
    side = Side(style="thin", color=theme.border_color)
    return Border(left=side, right=side, top=side, bottom=side)


def _source_for(group_name: str, config: PipelineConfig) -> str:
    """Data-dictionary 'Source' value for a column's group."""
    if group_name in config.taxonomies:
        return group_name
    return {"Identity": "incidents.csv", "Coverage": "Derived"}.get(group_name, "-")


def _band_label(group_name: str, config: PipelineConfig) -> str:
    """Return the band label for a group, falling back to the group name."""
    if group_name in config.styles:
        return config.styles[group_name].band_label
    if group_name in config.taxonomies:
        return config.taxonomies[group_name].band_label
    return group_name


def _style_body(ws: Worksheet, first_row: int, ncols: int, theme: ThemeConfig) -> None:
    """Apply alternating fills + thin borders + body font to all data rows."""
    border = _thin_border(theme)
    body_font = Font(name=theme.font, size=9)
    for r in range(first_row, ws.max_row + 1):
        fill = (
            PatternFill("solid", fgColor=theme.white)
            if r % 2 == 0
            else PatternFill("solid", fgColor=theme.alt_fill)
        )
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = body_font
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=False)


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _write_main_sheet(df: pd.DataFrame, writer: pd.ExcelWriter, config: PipelineConfig) -> None:
    """Write the Incidents sheet: title banner + group band + colored headers + striped body."""
    theme = config.theme
    df.to_excel(writer, sheet_name="Incidents", index=False, startrow=2)
    ws = writer.sheets["Incidents"]
    cols = list(df.columns)
    n = len(cols)
    last = get_column_letter(n)

    # Row 1 — title banner
    ws.merge_cells(f"A1:{last}1")
    cell = ws.cell(row=1, column=1)
    cell.value = f"AI Incident Database - Incidents | {len(df):,} incidents | {n} columns"
    cell.font = Font(name=theme.font, bold=True, size=11, color=theme.white)
    cell.fill = PatternFill("solid", fgColor=theme.banner_color)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Row 2 — group band (contiguous runs of the same group)
    current_group: str | None = None
    group_start = 1
    for idx, col in enumerate(cols, 1):
        group = _col_group(col, config)[0]
        if group != current_group:
            if current_group is not None:
                ws.merge_cells(start_row=2, start_column=group_start, end_row=2, end_column=idx - 1)
                cell = ws.cell(row=2, column=group_start)
                cell.value = _band_label(current_group, config)
                cell.font = Font(name=theme.font, bold=True, size=8, color=theme.white)
                cell.fill = PatternFill("solid", fgColor=_col_group(cols[group_start - 1], config)[1])
                cell.alignment = Alignment(horizontal="left", vertical="center")
            current_group, group_start = group, idx
    # flush last run
    ws.merge_cells(start_row=2, start_column=group_start, end_row=2, end_column=n)
    cell = ws.cell(row=2, column=group_start)
    cell.value = _band_label(current_group, config)
    cell.font = Font(name=theme.font, bold=True, size=8, color=theme.white)
    cell.fill = PatternFill("solid", fgColor=_col_group(cols[group_start - 1], config)[1])
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    # Row 3 — column headers
    border = _thin_border(theme)
    for idx, col in enumerate(cols, 1):
        _, hex_color = _col_group(col, config)
        cell = ws.cell(row=3, column=idx)
        cell.font = Font(name=theme.font, bold=True, size=9, color=theme.white)
        cell.fill = PatternFill("solid", fgColor=hex_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[3].height = 36

    # Body rows
    _style_body(ws, 4, n, theme)

    # Column widths
    for idx, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(idx)].width = config.output.column_widths.get(col, 18)

    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:{last}3"


def _write_table_sheet(name: str, df: pd.DataFrame, writer: pd.ExcelWriter, config: PipelineConfig) -> None:
    """Write Reports or Entities: title banner + navy header + striped body."""
    theme = config.theme
    df.to_excel(writer, sheet_name=name, index=False, startrow=1)
    ws = writer.sheets[name]
    n = len(df.columns)
    last = get_column_letter(n)

    # Row 1 — banner
    ws.merge_cells(f"A1:{last}1")
    cell = ws.cell(row=1, column=1)
    cell.value = f"AI Incident Database - {name} | {len(df):,} {name.lower()}"
    cell.font = Font(name=theme.font, bold=True, size=11, color=theme.white)
    cell.fill = PatternFill("solid", fgColor=theme.banner_color)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Row 2 — headers
    border = _thin_border(theme)
    for idx in range(1, n + 1):
        cell = ws.cell(row=2, column=idx)
        cell.font = Font(name=theme.font, bold=True, size=9, color=theme.white)
        cell.fill = PatternFill("solid", fgColor=theme.generic_header)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[2].height = 20

    # Body rows
    _style_body(ws, 3, n, theme)

    # Column widths
    for idx, col in enumerate(df.columns, 1):
        width = 50 if col.lower() in ("description", "title", "url", "text") else 20
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{last}2"


def _write_dictionary_sheet(df: pd.DataFrame, writer: pd.ExcelWriter, config: PipelineConfig) -> None:
    """Write the Data Dictionary sheet: 5 columns + banner."""
    theme = config.theme
    rows = []
    for col in df.columns:
        group_name, _ = _col_group(col, config)
        rows.append({
            "Column": col,
            "Group": group_name,
            "Source": _source_for(group_name, config),
            "Fill Rate": f"{df[col].notna().mean() * 100:.0f}%",
            "Description": config.column_descriptions.get(col, "-"),
        })
    dict_df = pd.DataFrame(rows, columns=["Column", "Group", "Source", "Fill Rate", "Description"])
    dict_df.to_excel(writer, sheet_name="Data Dictionary", index=False, startrow=1)

    ws = writer.sheets["Data Dictionary"]

    # Row 1 — banner (A1:E1)
    ws.merge_cells("A1:E1")
    cell = ws.cell(row=1, column=1)
    cell.value = "Data Dictionary - AI Incident Database"
    cell.font = Font(name=theme.font, bold=True, size=12, color=theme.white)
    cell.fill = PatternFill("solid", fgColor=theme.banner_color)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    # Row 2 — headers
    border = _thin_border(theme)
    for idx in range(1, 6):
        cell = ws.cell(row=2, column=idx)
        cell.font = Font(name=theme.font, bold=True, size=9, color=theme.white)
        cell.fill = PatternFill("solid", fgColor=theme.generic_header)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Body rows — inline to allow Group-cell coloring
    body_font = Font(name=theme.font, size=9)
    for r in range(3, ws.max_row + 1):
        col_name_cell = ws.cell(row=r, column=1).value  # Column name for group lookup
        _, grp_hex = _col_group(str(col_name_cell) if col_name_cell else "", config)
        fill = (
            PatternFill("solid", fgColor=theme.alt_fill)
            if r % 2 == 0
            else PatternFill("solid", fgColor=theme.white)
        )
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if c == 2:
                # Group column: bold, colored per group
                cell.font = Font(name=theme.font, size=9, bold=True, color=grp_hex)
            else:
                cell.font = body_font

    # Column widths
    for col_letter, width in zip(["A", "B", "C", "D", "E"], [28, 12, 14, 10, 72]):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A3"


def _write_coverage_sheet(df: pd.DataFrame, writer: pd.ExcelWriter, config: PipelineConfig) -> None:
    """Write the Coverage Map sheet: 4 columns + TOTAL row."""
    if "Data Sources" not in df.columns:
        return

    theme = config.theme
    total = len(df)
    breakdown = df["Data Sources"].value_counts()

    cov_rows = [
        [val, cnt, f"{cnt / total * 100:.1f}%", config.coverage_analysis.get(val, "-")]
        for val, cnt in breakdown.items()
    ]
    cov_rows.append(["TOTAL", total, "100%", ""])
    cov_df = pd.DataFrame(
        cov_rows,
        columns=["Data Sources", "Incidents", "% of Total", "What you can analyze"],
    )
    cov_df.to_excel(writer, sheet_name="Coverage Map", index=False, startrow=1)

    ws = writer.sheets["Coverage Map"]

    # Row 1 — banner (A1:D1)
    ws.merge_cells("A1:D1")
    cell = ws.cell(row=1, column=1)
    cell.value = "Coverage Map - What you can analyze at each taxonomy level"
    cell.font = Font(name=theme.font, bold=True, size=11, color=theme.white)
    cell.fill = PatternFill("solid", fgColor=theme.banner_color)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Row 2 — headers
    border = _thin_border(theme)
    for idx in range(1, 5):
        cell = ws.cell(row=2, column=idx)
        cell.font = Font(name=theme.font, bold=True, size=9, color=theme.white)
        cell.fill = PatternFill("solid", fgColor=theme.generic_header)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Body rows (stripes + borders)
    _style_body(ws, 3, 4, theme)

    # Override TOTAL row
    total_row = ws.max_row
    total_fill = PatternFill("solid", fgColor=theme.total_fill)
    total_font = Font(name=theme.font, bold=True, size=9)
    for c in range(1, 5):
        cell = ws.cell(row=total_row, column=c)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border

    # Column widths
    for col_letter, width in zip(["A", "B", "C", "D"], [24, 14, 12, 68]):
        ws.column_dimensions[col_letter].width = width


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def export_excel(df: pd.DataFrame, reports: pd.DataFrame, entities: pd.DataFrame, config: PipelineConfig) -> None:
    """Write the Excel export workbook with multiple sheets."""
    output_path = config.paths.output_path
    output_path.parent.mkdir(parents=True, exist_ok=True)

    df = _apply_order(df, config.output.column_order)
    reports = _apply_order(reports, config.output.reports_column_order)
    entities = _apply_order(entities, config.output.entities_column_order)

    # Clean timezone data if any timestamps are tz-aware (Excel doesn't support them directly).
    for d in [df, reports, entities]:
        for col in d.select_dtypes(include=["datetime64[ns, UTC]"]).columns:
            d[col] = d[col].dt.tz_localize(None)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _write_main_sheet(df, writer, config)
        _write_table_sheet("Reports", reports, writer, config)
        _write_table_sheet("Entities", entities, writer, config)
        _write_dictionary_sheet(df, writer, config)
        _write_coverage_sheet(df, writer, config)
